from utilities.hash_table import HashTable
from typing import cast, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

class Package:
    def __init__(self, package_id: int, address: str, deadline: str, city: str, zip_code: str, weight: int, note: str):
        self.package_id = package_id
        self.address = address
        self.deadline = deadline
        self.city = city
        self.zip_code = zip_code
        self.weight = weight
        self.note = note
        self.status = "at the hub"

class PackageTable:
    def __init__(self, package_file_path: str):
        self._package_table: HashTable[Package] = HashTable()
        self._load_package_data(package_file_path)

    def get_package(self, package_id: int) -> Optional[Package]:
        return self._package_table.get(package_id)

    def _load_package_data(self, package_file_path: str):
        wb = load_workbook(package_file_path)

        if wb is None:
            raise Exception("Package file not found")

        ws = cast( Worksheet, wb.active )

        for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=8):

            row_values = []
            for cell in row:
                row_values.append(cell.value)

            package_id = int(row_values[0])
            address = row_values[1]
            deadline = row_values[5]
            city = row_values[2]
            zip_code = row_values[4]
            weight = int(row_values[6])
            note = row_values[7]

            package = Package(package_id, address, deadline, city, zip_code, weight, note)

            self._package_table.insert(package_id, package)
        
